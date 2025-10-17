import sys,openpyxl, os , os.path ,datetime ,re, time , shutil
from openpyxl import load_workbook

# Get folder where the exe is running
if getattr(sys, 'frozen', False):  
    # Running as exe
    script_dir = os.path.dirname(sys.executable)
else:  
    # Running as .py
    script_dir = os.path.dirname(os.path.abspath(__file__))

# catch all files
script_dir = os.path.dirname(os.path.abspath(__file__))
files = os.listdir(script_dir)

# ---
# create report free files directory for moving
os.makedirs(os.path.join(script_dir, "nonReportFiles"),exist_ok=True)

#---
# calculate last 3 working days (code from chatGpt)
def is_last3_workdays(date_str):
    from datetime import datetime, timedelta
    # Parse the date string
    date = datetime.strptime(date_str, "%d-%m-%Y").date()
    year, month = date.year, date.month

    # Find last day of month
    if month == 12:
        next_month = datetime(year + 1, 1, 1).date()
    else:
        next_month = datetime(year, month + 1, 1).date()
    last_day = next_month - timedelta(days=1)

    # Collect last 3 working days (Sun–Thu)
    workdays = []
    d = last_day
    while len(workdays) < 3:
        if d.weekday() not in (4, 5):  # Friday=4, Saturday=5
            workdays.append(d)
        d -= timedelta(days=1)

    return date in workdays
# ---

def overCelling(data_list):
    # print ("Overcelling:")
    tmpAm=[]
    for i in range (len(data_list)):
        if len(tmpAm)==2:
            break
        if "Acc No:" in data_list[i][0]:
            match = re.search(r"Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[i][0])
            tmpAm.append(float(match.group(1).replace(",","")))
    if len(tmpAm)==2 and tmpAm[0] > tmpAm[1]*40/100+tmpAm[1]:
        # print (data_list[6])
        reports[0].append(fileName)
        global reportCheck
        reportCheck=1

def overAdvance(data_list):
    # print ("OverAdvance:")
    tmpAm=0
    count=0
    for i in range (0,len(data_list)):
        if "Acc No:" in data_list[i][0]:
            match = re.search(r"Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[i][0])
            tmpAm=float(match.group(1).replace(",",""))
        if "Member Total" in data_list[i][0]:
            count+=1
            if count == 2:
                if tmpAm <600000:
                    inst=tmpAm*0.095
                elif tmpAm <1000000:
                    inst=tmpAm*0.094
                else:
                    inst=tmpAm*0.093
                coll=data_list[i-1][3]
                if coll>inst*3:
                    # print (data_list[6])
                    reports[1].append(fileName)
                    global reportCheck
                    reportCheck=1
                break

def overDuePreviousLoan(data_list):
    count=0
    for i in range (len(data_list)):
        if "Acc No:" in data_list[i][0]:
            count+=1
            if count!=1: # capturing the 2nd loan
                # data clean , replace comma
                for x in range (i+2,len(data_list)):
                    if "Member Total" in data_list[x][0]: # end of the loan data
                        break
                    data_list[x][3]=float(str(data_list[x][3]).replace(",", "").strip()) if data_list[x][3] not in (None, "") else 0
                # ---
                targetAm=0.00
                collAm=0.00
                inst=0.00
                firstCollDate=data_list[i+2][0]
                match = re.search(r"Disb\. Date:\s*([\d-]+).*?Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[i][0])
                tmpAm=float(match.group(2).replace(",","")) # remove comma from Disburse Amount
                if tmpAm <500000:
                    inst=tmpAm*0.095
                elif tmpAm <1000000:
                    inst=tmpAm*0.094
                else:
                    inst=tmpAm*0.093
                j=i+2
                while j <len(data_list):
                    # print(data_list[j])
                    if "Member Total" in data_list[j][0]: # end of the loan data
                        break
                    # print(data_list[j][0])
                    # some tran: has N/A value , needs to clean with 0
                    if data_list[j][2]=="N/A":
                        data_list[j][2]=0.00
                    # ---
                    # if same month/year found , collection add
                    if firstCollDate.split("-")[1]==data_list[j][0].split("-")[1] and firstCollDate.split("-")[2]==data_list[j][0].split("-")[2]:
                        # if disb and collection same month, target amount should reduce by collection amount
                        if match.group(1).split("-")[1]==firstCollDate.split("-")[1] and match.group(1).split("-")[2]==firstCollDate.split("-")[2]:
                            targetAm-=data_list[j][3]
                        collAm+=data_list[j][3]
                    else:
                        targetAm+=inst # if month change , then target increase by installment
                        firstCollDate=data_list[j][0] # then month swipe
                        if targetAm>collAm: # if target amount bigger then collection amount, then it is od
                            # print(allFiles)
                            # print ("od",targetAm,collAm)
                            reports[2].append(fileName)
                            global reportCheck
                            reportCheck=1
                            return
                        else:
                            # j should not be increase because month may missed
                            continue
                    j=j+1

def disburseInstallmentSameMonth(data_list):
    # print ("Same Month Disbursement and Installment:")
    match = re.search(r"Disb\. Date:\s*([\d-]+).*?Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[7][0])
    for i in range (9,len(data_list)):
        if data_list[i][3]>0:
            if data_list[i][0].split("-")[1]==match.group(1).split("-")[1]:
                # print(data_list[6])
                reports[5].append(fileName)
                global reportCheck
                reportCheck=1
        break

def installmentDisburseSameDate (data_list):
    # print ("Same Date Installment and Disburse:")
    count=0
    collDate=""
    for i in range (len(data_list)):
        if "Acc No:" in data_list[i][0]:
            count+=1
            if count==2: # capturing the 2nd loan
                # data clean , replace comma
                for x in range (i+2,len(data_list)):
                    if "Member Total" in data_list[x][0]: # end of the loan data
                        collDate=data_list[x-1][0]
                        break
    match = re.search(r"Disb\. Date:\s*([\d-]+).*?Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[7][0])
    if match.group(1).split("-")[0]==collDate.split("-")[0] and match.group(1).split("-")[1]==collDate.split("-")[1] and match.group(1).split("-")[2]==collDate.split("-")[2]:
        # print(allFiles)
        reports[4].append(fileName)
        global reportCheck
        reportCheck=1
    return

def bulkAdvanceRealization_keepBorrowerSmallAmountOutstanding (data_list):
    for i in range (len(data_list)):
        if "Acc No:" in data_list[i][0]:
            match = re.search(r"Disb\. Date:\s*([\d-]+).*?Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[i][0])
            tmpAm=float(match.group(2).replace(",","")) # remove comma from Disburse Amount
            inst=0
            if tmpAm <500000:
                inst=tmpAm*0.095
            elif tmpAm <1000000:
                inst=tmpAm*0.094
            else:
                inst=tmpAm*0.093
            for x in range (i+2,len(data_list)):
                # print(data_list[x][3])
                if "Member Total" in data_list[x][0]: # end of the loan data
                    return
                if data_list[x][3]>inst*3:
                    reports[6].append(fileName)
                    global reportCheck
                    reportCheck=1
                    return


def disburseAfter20 (data_list):
        match = re.search(r"Disb\. Date:\s*([\d-]+).*?Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[7][0])
        if int(match.group(1).split("-")[0])>20:
            reports[7].append(fileName)
            global reportCheck
            reportCheck=1

def RealizationLast3WorkingDaysPreviousLoan (data_list):
    # doesn't work if wrong posting happen from accounts within targeted days (need manual check)
    count=0
    for i in range (len(data_list)):
        if "Acc No:" in data_list[i][0]:
            # print(data_list[i][0])
            count+=1
            if count==2: # capturing the 2nd loan
                # data clean , replace comma
                for x in range (i+2,len(data_list)):
                    if "Member Total" in data_list[x][0]: # end of the loan data
                        break
                    data_list[x][3]=float(str(data_list[x][3]).replace(",", "").strip()) if data_list[x][3] not in (None, "") else 0
                    # print(data_list[x][3])
                # ---
                targetAm=0.00
                collAm=0.00
                inst=0.00
                firstCollDate=data_list[i+2][0]
                match = re.search(r"Disb\. Date:\s*([\d-]+).*?Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[i][0])
                tmpAm=float(match.group(2).replace(",","")) # remove comma from Disburse Amount
                if tmpAm <500000:
                    inst=tmpAm*0.095
                elif tmpAm <1000000:
                    inst=tmpAm*0.094
                else:
                    inst=tmpAm*0.093
                j=i+2
                while j <len(data_list):
                    # print(data_list[j])
                    if "Member Total" in data_list[j][0]: # end of the loan data
                        break
                    # print(data_list[j][0])
                    # some tran: has N/A value , needs to clean with 0
                    if data_list[j][2]=="N/A":
                        data_list[j][2]=0.00
                    # ---
                    # if same month/year found , collection add
                    flag=0
                    if firstCollDate.split("-")[1]==data_list[j][0].split("-")[1] and firstCollDate.split("-")[2]==data_list[j][0].split("-")[2]:
                        # if disburse and collection same month, target amount should reduce by collection amount
                        if match.group(1).split("-")[1]==firstCollDate.split("-")[1] and match.group(1).split("-")[2]==firstCollDate.split("-")[2]:
                            targetAm-=data_list[j][3]
                        collAm+=data_list[j][3]
                        if data_list[j][3] > 0:
                            # lastCollDate=int(data_list[j][0].split("-")[0])
                            lastCollDate=data_list[j][0]
                            flag=1
                            # print(data_list[j][0])
                        if flag==0:
                            # lastCollDate=int(data_list[j][0].split("-")[0])
                            lastCollDate=data_list[j][0]

                    else:
                        # print(lastCollDate)
                        targetAm+=inst # if month change , then target increased by installment
                        firstCollDate=data_list[j][0] # then month swipe
                        # collAm==targetAm collAm<targetAm+inst collAm < targetAm
                        if (targetAm+inst>collAm and targetAm<=collAm) and is_last3_workdays(lastCollDate): # and (lastCollDate == 31 or lastCollDate == 30 or lastCollDate == 29 or lastCollDate == 28 or lastCollDate == 27 or lastCollDate == 26) :
                            # if is_last3_workdays(lastCollDate):
                            # print(allFiles)
                            # print ("Last 3 Working Days Installment Previous Loan",targetAm,collAm,data_list[j][0],lastCollDate)
                            reports[3].append(fileName)
                            global reportCheck
                            reportCheck=1
                            return
                        else:
                            # j should not be increase because month may missed
                            continue
                    j=j+1

def RealizationLast3WorkingDaysCurrentLoan (data_list):
    # doesn't work if wrong posting happen from accounts within targeted days (need manual check)
    count=0
    for i in range (len(data_list)):
        if "Acc No:" in data_list[i][0]:
            # print(data_list[i][0])
            count+=1
            if count==1: # capturing the 1st loan
                # data clean , replace comma
                for x in range (i+2,len(data_list)):
                    if "Member Total" in data_list[x][0]: # end of the loan data
                        if data_list[x-1][6]==0: # check if loan is cleared
                            # print (data_list[x-1][6])
                            return
                        break
                    data_list[x][3]=float(str(data_list[x][3]).replace(",", "").strip()) if data_list[x][3] not in (None, "") else 0
                    # print(data_list[x][3])
                # ---
                targetAm=0.00
                collAm=0.00
                inst=0.00
                firstCollDate=data_list[i+2][0]
                match = re.search(r"Disb\. Date:\s*([\d-]+).*?Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[i][0])
                tmpAm=float(match.group(2).replace(",","")) # remove comma from Disburse Amount
                if tmpAm <500000:
                    inst=tmpAm*0.095
                elif tmpAm <1000000:
                    inst=tmpAm*0.094
                else:
                    inst=tmpAm*0.093
                j=i+2
                while j <len(data_list):
                    # print(data_list[j])
                    if "Member Total" in data_list[j][0]: # end of the loan data
                        break
                    # print(data_list[j][0])
                    # some tran: has N/A value , needs to clean with 0
                    if data_list[j][2]=="N/A":
                        data_list[j][2]=0.00
                    # ---
                    # if same month/year found , collection add
                    flag=0
                    if firstCollDate.split("-")[1]==data_list[j][0].split("-")[1] and firstCollDate.split("-")[2]==data_list[j][0].split("-")[2]:
                        # if disburse and collection same month, target amount should reduce by collection amount
                        if match.group(1).split("-")[1]==firstCollDate.split("-")[1] and match.group(1).split("-")[2]==firstCollDate.split("-")[2]:
                            targetAm-=data_list[j][3]
                        collAm+=data_list[j][3]
                        if data_list[j][3] > 0:
                            # lastCollDate=int(data_list[j][0].split("-")[0])
                            lastCollDate=data_list[j][0]
                            flag=1
                            # print(data_list[j][0])
                        if flag==0:
                            # lastCollDate=int(data_list[j][0].split("-")[0])
                            lastCollDate=data_list[j][0]
                    else:
                        # print(lastCollDate)
                        targetAm+=inst # if month change , then target increased by installment
                        firstCollDate=data_list[j][0] # then month swipe
                        # collAm==targetAm collAm<targetAm+inst collAm < targetAm
                        if (targetAm+inst>collAm and targetAm<=collAm) and is_last3_workdays(lastCollDate): # and (lastCollDate == 31 or lastCollDate == 30 or lastCollDate == 29 or lastCollDate == 28 or lastCollDate == 27 or lastCollDate == 26) :
                            # if is_last3_workdays(lastCollDate):
                            # print(allFiles)
                            # print ("Last 3 Working Days Installment Current Loan",targetAm,collAm,data_list[j][0],lastCollDate)
                            reports[8].append(fileName)
                            global reportCheck
                            reportCheck=1
                            return
                        else:
                            # j should not be increase because month may missed
                            continue
                    j=j+1

def brokenInstallmentPreviousLoan (data_list):
    # doesn't work if wrong posting happen form account bcz collection amount will adjust & broken inst value will increase
    count=0
    for i in range (len(data_list)):
        if "Acc No:" in data_list[i][0]:
            # print(data_list[i][0])
            count+=1
            if count==2: # capturing the 2nd loan
                # data clean , replace comma
                for x in range (i+2,len(data_list)):
                    if "Member Total" in data_list[x][0]: # end of the loan data
                        break
                    data_list[x][3]=float(str(data_list[x][3]).replace(",", "").strip()) if data_list[x][3] not in (None, "") else 0
                    # print(data_list[x][3])
                # ---
                targetAm=0.00
                collAm=0.00
                inst=0.00
                brokenIsnt=0
                firstCollDate=data_list[i+2][0]
                match = re.search(r"Disb\. Date:\s*([\d-]+).*?Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[i][0])
                tmpAm=float(match.group(2).replace(",","")) # remove comma from Disburse Amount
                if tmpAm <500000:
                    inst=tmpAm*0.095
                elif tmpAm <1000000:
                    inst=tmpAm*0.094
                else:
                    inst=tmpAm*0.093
                j=i+2
                while j <len(data_list):
                    # print(data_list[j])
                    if "Member Total" in data_list[j][0]: # end of the loan data
                        break
                    # print(data_list[j][0])
                    # some tran: has N/A value , needs to clean with 0
                    if data_list[j][2]=="N/A":
                        data_list[j][2]=0.00
                    # ---
                    # if same month/year found , collection add
                    flag=0
                    if firstCollDate.split("-")[1]==data_list[j][0].split("-")[1] and firstCollDate.split("-")[2]==data_list[j][0].split("-")[2]:
                        # if disburse and collection same month, target amount should reduce by collection amount
                        if match.group(1).split("-")[1]==firstCollDate.split("-")[1] and match.group(1).split("-")[2]==firstCollDate.split("-")[2]:
                            return
                            targetAm-=data_list[j][3]
                        collAm+=data_list[j][3]
                        if data_list[j][3] > 0:
                            brokenIsnt+=1
                            lastCollDate=data_list[j][0]
                            flag=1
                        if flag==0:
                            lastCollDate=data_list[j][0]

                    else:
                        targetAm+=inst # if month change , then target increased by installment
                        firstCollDate=data_list[j][0] # then month swipe
                        if (targetAm+1500>collAm and targetAm<=collAm) and brokenIsnt>1:
                            reports[9].append(fileName)
                            global reportCheck
                            reportCheck=1
                            return
                        else:
                            # j should not be increase because month may missed
                            brokenIsnt=0
                            continue
                    j=j+1

def brokenInstallmentCurrentLoan (data_list):
    count=0
    for i in range (len(data_list)):
        if "Acc No:" in data_list[i][0]:
            # print(data_list[i][0])
            count+=1
            if count==1: # capturing the 1st loan
                # data clean , replace comma
                for x in range (i+2,len(data_list)):
                    if "Member Total" in data_list[x][0]: # end of the loan data
                        break
                    data_list[x][3]=float(str(data_list[x][3]).replace(",", "").strip()) if data_list[x][3] not in (None, "") else 0
                    # print(data_list[x][3])
                # ---
                targetAm=0.00
                collAm=0.00
                inst=0.00
                brokenIsnt=0
                firstCollDate=data_list[i+2][0]
                match = re.search(r"Disb\. Date:\s*([\d-]+).*?Disb\. Amount:\s*([\d,]+\.\d{2})", data_list[i][0])
                tmpAm=float(match.group(2).replace(",","")) # remove comma from Disburse Amount
                if tmpAm <500000:
                    inst=tmpAm*0.095
                elif tmpAm <1000000:
                    inst=tmpAm*0.094
                else:
                    inst=tmpAm*0.093
                j=i+2
                while j <len(data_list):
                    # print(data_list[j])
                    if "Member Total" in data_list[j][0]: # end of the loan data
                        break
                    # print(data_list[j][0])
                    # some tran: has N/A value , needs to clean with 0
                    if data_list[j][2]=="N/A":
                        data_list[j][2]=0.00
                    # ---
                    # if same month/year found , collection add
                    flag=0
                    if firstCollDate.split("-")[1]==data_list[j][0].split("-")[1] and firstCollDate.split("-")[2]==data_list[j][0].split("-")[2]:
                        # if disburse and collection same month, target amount should reduce by collection amount
                        if match.group(1).split("-")[1]==firstCollDate.split("-")[1] and match.group(1).split("-")[2]==firstCollDate.split("-")[2]:
                            return
                            targetAm-=data_list[j][3]
                        collAm+=data_list[j][3]
                        if data_list[j][3] > 0:
                            brokenIsnt+=1
                            lastCollDate=data_list[j][0]
                            flag=1
                        if flag==0:
                            lastCollDate=data_list[j][0]

                    else:
                        targetAm+=inst # if month change , then target increased by installment
                        firstCollDate=data_list[j][0] # then month swipe
                        if (targetAm+1500>collAm and targetAm<=collAm) and brokenIsnt>1:
                            reports[10].append(fileName)
                            global reportCheck
                            reportCheck=1
                            return
                        else:
                            # j should not be increase because month may missed
                            brokenIsnt=0
                            continue
                    j=j+1
            break



reports=[
    ["overCelling:"],
    ["overAdvance:"],
    ["overDuePreviousLoan:"],
    ["RealizationLast3WorkingDaysPreviousLoan:"],
    ["installmentDisburseSameDate:"],
    ["disburseInstallmentSameMonth:"],
    ["bulkAdvanceRealization_keepBorrowerSmallAmountOutstanding:"],
    ["disburseAfter20:"],
    ["RealizationLast3WorkingDaysCurrentLoan:"],
    ["brokenInstallmentPreviousLoan:"],
    ["brokenInstallmentCurrentLoan:"],
]
fileName=""
fileCount=0
reportCheck=0
closeFileCount=0
nonReportFilesCount=0
for allFiles in files:
    if allFiles.startswith("SB-") and allFiles.endswith(".xlsx"):
        fileCount+=1
        # print (allFiles)

        # Load an existing Excel file   SB-LOANCOLLECTION_MEMBERWISE_19092025_032023PM  SB-LOANCOLLECTION_MEMBERWISE_19092025_031758PM SB-LOANCOLLECTION_MEMBERWISE_21092025_010900PM.pdf SB-LOANCOLLECTION_MEMBERWISE_20092025_032144AM
        # wb = load_workbook(os.path.join(script_dir,"SB-LOANCOLLECTION_MEMBERWISE_19092025_102046PM.xlsx"), data_only=True)  # data_only=True → read values instead of formulas
        wb = load_workbook(os.path.join(script_dir,allFiles), data_only=True)  # data_only=True → read values instead of formulas
        ws = wb.active  # get active sheet
        # grabbing all data form excel file 
        data_list = [[
            cell.strftime("%d-%m-%Y") if isinstance(cell, datetime.datetime) else (cell if cell is not None else "")
            for cell in row
        ] for row in ws.iter_rows(values_only=True)]
        # ---
        reportCheck=0
        # check if loan is closed 
        closeFileCheck=0
        for i in range (len(data_list)):
            if "Acc No:" in data_list[i][0]:
                for x in range (i+2,len(data_list)):
                    if "Member Total" in data_list[x][0]: # end of the loan data
                        if data_list[x-1][8]==float(0):
                            # print ("catched",data_list[x-1][8],fileName)
                            closeFileCheck=1
                        break
                break
        if closeFileCheck==1:
            closeFileCount+=1
            shutil.move(os.path.join(script_dir, allFiles),os.path.join(script_dir, "nonReportFiles"))
            # shutil.move("nonReportFiles",allFiles)
            continue
        # ---

        os.rename(os.path.join(script_dir, allFiles),os.path.join(script_dir, "processing_"+allFiles))
        Project=data_list[4][0][data_list[4][0].index('[')+2:data_list[4][0].index(']')-1]
        Branch=data_list[5][0][data_list[5][0].index('h:')+3:data_list[5][0].index(']')-1].replace("["," ").replace("/"," ")
        Member=data_list[6][0][data_list[6][0].index('er:')+3:data_list[6][0].rindex('[')-1]
        Member2=data_list[6][0][data_list[6][0].index('er:')+3:data_list[6][0].rindex(']')+1]
        MemberId=Member2[Member2.index('[')+1:Member2.index(']')]
        Po=data_list[6][0][data_list[6][0].index('PO:')+4:data_list[6][0].index('[')-1]
        DisbDate=data_list[7][0][data_list[7][0].index('te:')+4:data_list[7][0].index('| Disb. A')-1].replace("-",".")
        DisbAm=data_list[7][0][data_list[7][0].index('nt:')+4:data_list[7][0].index('| Loan S')-1]

        fileName = "_".join((MemberId + "-" + Member + "-" + Project + "-" + DisbDate + "-" + DisbAm + "-" + Po + "-" + Branch).split())

        overCelling(data_list) #0
        overAdvance(data_list) #1
        overDuePreviousLoan(data_list) #2
        RealizationLast3WorkingDaysPreviousLoan (data_list) #3
        installmentDisburseSameDate(data_list) #4
        disburseInstallmentSameMonth(data_list) #5
        bulkAdvanceRealization_keepBorrowerSmallAmountOutstanding (data_list) #6
        disburseAfter20 (data_list) #7
        RealizationLast3WorkingDaysCurrentLoan (data_list) #problem in SB-LOANCOLLECTION_MEMBERWISE_19092025_102153PM #8
        brokenInstallmentPreviousLoan (data_list) #9
        brokenInstallmentCurrentLoan (data_list) #10
        # script_dir = os.path.dirname(os.path.abspath(__file__))
        # print(text)
        print(allFiles,fileName+".xlsx")
        # time.sleep(2)
        os.rename(os.path.join(script_dir, "processing_"+allFiles),os.path.join(script_dir, fileName+".xlsx"))
        
        # Moving normal files into the directory
        if reportCheck==0:
            nonReportFilesCount+=1
            shutil.move(os.path.join(script_dir, fileName+".xlsx"),os.path.join(script_dir, "nonReportFiles"))
        # ---
        # break


# for r in range (len(reports)):
#     print("")
#     for l in range(len(reports[r])):
#         None
#         print(reports[r][l])



with open(os.path.join(script_dir,"reports.txt"), "w") as file:
    for r in reports:
        print("")  # prints a blank line to console
        file.write("\n")  # writes a blank line to file
        for l in r:
            print(l)                # print to console
            file.write(str(l) + "\n")  # write to file

    print("\nFinished...\n","Total",fileCount,"Files checked.","Report Files:",fileCount-closeFileCount-nonReportFilesCount,", Non Report Files:",nonReportFilesCount,", Closed Files:",closeFileCount)
    file.write(f"\nFinished...\nTotal {fileCount} Files checked. Report Files: {fileCount - closeFileCount - nonReportFilesCount}, Non Report Files: {nonReportFilesCount}, Closed Files: {closeFileCount}\n")



# from reportlab.pdfgen import canvas
# import os

# # Optional: create folder to save PDF
# folder = "pdfReports"
# os.makedirs(folder, exist_ok=True)

# pdf_path = os.path.join(folder, "reportFile.pdf")
# pdf = canvas.Canvas(pdf_path)

# y = 750  # starting y-position on page

# for r in range(len(reports)):
#     print("")  # your original blank line, untouched
#     for l in range(len(reports[r])):
#         pdf.drawString(100, y, reports[r][l])  # write to PDF
#         print(reports[r][l])                   # your girl, untouched 😏
#         y -= 20  # move down for next line

#         # Start a new page if we reach bottom
#         if y < 50:
#             pdf.showPage()
#             y = 750

# pdf.save()
# print("PDF saved successfully!")

# x=input("Press any key...")